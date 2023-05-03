from django.shortcuts import render, redirect
from django.views.decorators import gzip
from django.http import StreamingHttpResponse
from django.urls import reverse
from django.core import serializers
from django.utils import timezone
from openpyxl import load_workbook
from .modules.camera import VideoCamera, gen
from .modules.forms import CreateCamera_form
from .models import Camera
import csv
import io

def cameras(request):
    cameras = Camera.objects.filter(deleted__isnull=True)
    context = {
        'cameras_list': cameras,
        'headerTitle': "Cámaras",
        'breadcrumb': [
            {'tag': 'Cámaras', 'url': 'cameras'}
        ],
        'form': CreateCamera_form()
    }
    if request.method == 'GET':
        context["show_alert"] = request.GET.get('show_alert', '')
        context["message"] = request.GET.get('message', '')
    if request.method == 'POST':
        if request.POST.get('upload_excel'):
            file=request.FILES['excel_cameras']
            camerasImported_list = get_cameras_direct_list(file)
            camerasFailed = []
            for cameraData in camerasImported_list:
                if 'camaras' in cameraData:
                    pass
                else:
                    try:
                        if cameraData[0] == '':
                            cameraData[0] = None
                        if cameraData[1] == '':
                            cameraData[1] = None
                        Camera.objects.create(
                            name = cameraData[0],
                            rtsp = cameraData[1],
                            peop_c_service = cameraData[2] == 'si',
                            face_rec_service = cameraData[3] == 'si',
                            vehicles_service = cameraData[4] == 'si',
                        )
                    except Exception as e:
                        camerasFailed.append(cameraData)
            if camerasFailed:
                context['show_alert'] = 'error'
                context['message'] = 'Error al importar archivo'
                context['camerasFailed'] = camerasFailed
            else:
                context['show_alert'] = 'success'
                context['message'] = 'Archivos importados desde el excel correctamente'
            return render(request, 'cameras_admin/cameras.html', context)        
    return render(request, 'cameras_admin/cameras.html', context)
    
def create_camera(request):
    if request.method == 'POST':
        form = CreateCamera_form(request.POST)
        if form.is_valid():
            form.save()
            cameras = Camera.objects.filter(deleted__isnull=True)
            message = "Cámara agregada correctamente"
            show_alert = "success"
            url = reverse('cameras') + f"?message={message}&show_alert={show_alert}"
            return redirect (url)
        else:
            cameras = Camera.objects.filter(deleted__isnull=True)
            context = {
                'headerTitle': "Cámaras",
                'breadcrumb': [
                    {'tag': 'Cámaras', 'url': 'cameras'}
                ],
                'form': form,
                'cameras': cameras,
            }
            return render(request, 'cameras_admin/cameras.html', context)
    else:
        form = CreateCamera_form()
        cameras = Camera.objects.filter(deleted__isnull=True)
        context = {
            'headerTitle': "Cámaras",
            'breadcrumb': [
                {'tag': 'Cámaras', 'url': 'cameras'}
            ],
            'form': form,
            'cameras': cameras,
        }
        return render(request, 'cameras_admin/cameras.html', context)
    
def edit_camera(request, id):
    cameras = Camera.objects.filter(deleted__isnull=True)
    context = {
        'cameras_list': cameras,
        'headerTitle': 'Cámaras',
        'breadcrumb': [
            {'tag': 'Cámaras', 'url': 'cameras'}
        ],
    }
    try:
        cameraById = Camera.objects.filter(pk=id, deleted__isnull=True).first()
        cameraById_json = serializers.serialize('json', [cameraById])
    except:
        return redirect('cameras')
    
    context['cameraById'] = cameraById
    context['cameraById_json'] = cameraById_json
    
    if request.method == 'POST':
        form = CreateCamera_form(request.POST)
        if form.is_valid():
            cameraById.name = request.POST.get('name')
            cameraById.rtsp = request.POST.get('rtsp')
            if(request.POST.get("peop_c_service")):
                cameraById.peop_c_service = True
            else:
                cameraById.peop_c_service = False
            if(request.POST.get("face_rec_service")):
                cameraById.face_rec_service = True
            else:
                cameraById.face_rec_service = False
            if(request.POST.get("vehicles_service")):
                cameraById.vehicles_service = True
            else:
                cameraById.vehicles_service = False
            cameraById.save()
            
            message = "Cámara editada correctamente"
            show_alert = "success"
            url = reverse('cameras') + f"?message={message}&show_alert={show_alert}"
            return redirect (url)
    return render(request, 'cameras_admin/cameras.html', context)

def delete_camera(request, id):
    cameras = Camera.objects.filter(deleted__isnull=True)
    context = {
        'cameras_list': cameras,
        'headerTitle': 'Cámaras',
        'breadcrumb': [
            {'tag': 'Cámaras', 'url': 'cameras'}
        ],
    }
    try:
        cameraById = Camera.objects.filter(pk=id, deleted__isnull=True).first()
    except:
        return redirect('cameras')
    
    context['cameraById'] = cameraById
    
    if request.method == 'POST':
        cameraById.deleted = timezone.now()
        cameraById.save()
        
        message = "Cámara eliminada correctamente"
        show_alert = "success"
        url = reverse('cameras') + f"?message={message}&show_alert={show_alert}"
        return redirect (url)
    
    return render(request, 'cameras_admin/cameras.html', context)

def rtsp_camera(request, id):
    try:
        cameraById = Camera.objects.filter(pk=id, deleted__isnull=True).first()
    except:
        return redirect('cameras')
    context = {
        'headerTitle': f"Video: {cameraById.name}",
        'breadcrumb': [
            {'tag': 'Cámaras', 'url': 'cameras'},
            {'tag': 'Video'}
        ],
        'camera_name': cameraById.name,
        'rtsp': cameraById.rtsp,
    }
    return render(request, 'cameras_admin/rtsp.html', context)

# GENERAR VIDEO POR RTSP
@gzip.gzip_page
def video_feed(request):
    rtsp = request.GET.get('rtsp', '')
    try:
        cam = VideoCamera(str(rtsp))
        return StreamingHttpResponse(gen(cam), content_type="multipart/x-mixed-replace;boundary=frame")
    except Exception as e:
        print(f"\nError \n'{e}'\n en 'video_feed'\n")

def get_cameras_direct_list(file):
    match file.name.split('.')[-1]:
        case 'xlsx':
            workbook = load_workbook(file)
            worksheet = workbook.active
            return list(worksheet.iter_rows(values_only=True))
        case 'csv':
            return csv.reader(io.TextIOWrapper(file.file, encoding='utf-8'))
        case _:
            return 'error'